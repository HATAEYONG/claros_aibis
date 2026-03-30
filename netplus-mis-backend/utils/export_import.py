# -*- coding: utf-8 -*-
"""
내보내기/가져오기를 위한 믹스인
"""
import csv
import io
import openpyxl
from rest_framework import status
from rest_framework.response import Response
from rest_framework.decorators import action
from django.http import HttpResponse
from datetime import datetime


class ExportImportMixin:
    """
    내보내기/가져오기 기능을 위한 믹스인 클래스

    지원 형식:
    - CSV (Comma Separated Values)
    - Excel (XLSX)
    - JSON
    """

    def get_export_fields(self):
        """
        내보내기 필드 목록을 반환합니다.
        ViewSet에서 오버라이드하여 사용자 정의 필드를 지정할 수 있습니다.
        """
        if hasattr(self, 'serializer_class') and self.serializer_class:
            serializer = self.serializer_class()
            return list(serializer.fields.keys())
        return []

    def get_export_field_names(self):
        """
        내보내기 필드의 한글 이름을 반환합니다.
        ViewSet에서 오버라이드하여 사용자 정의 이름을 지정할 수 있습니다.
        """
        return {}

    @action(detail=False, methods=['get'], url_path='export/csv')
    def export_csv(self, request):
        """
        CSV 형식으로 데이터 내보내기

        Query Parameters:
        - fields: 선택적 필드 목록 (쉼표로 구분)
        - filters: 필터 조건 (JSON 형식)
        - ordering: 정렬 필드

        Example: /api/production/work-orders/export/csv?fields=id,order_number,status
        """
        queryset = self.get_queryset()

        # 필터 적용
        queryset = self.filter_queryset(queryset)

        # 정렬 적용
        ordering = request.query_params.get('ordering')
        if ordering:
            queryset = queryset.order_by(ordering)

        # 필드 제한
        fields_param = request.query_params.get('fields')
        if fields_param:
            fields = fields_param.split(',')
        else:
            fields = self.get_export_fields()

        # 데이터 생성
        data = []
        for item in queryset:
            serializer = self.get_serializer(item)
            row = {k: str(v) if v is not None else '' for k, v in serializer.data.items() if k in fields}
            data.append(row)

        # CSV 생성
        response = HttpResponse(content_type='text/csv')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'{self.queryset.model._meta.model_name}_{timestamp}.csv'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.DictWriter(response, fieldnames=fields)
        writer.writeheader()
        writer.writerows(data)

        return response

    @action(detail=False, methods=['get'], url_path='export/excel')
    def export_excel(self, request):
        """
        Excel 형식으로 데이터 내보내기

        Query Parameters:
        - fields: 선택적 필드 목록 (쉼표로 구분)
        - filters: 필터 조건 (JSON 형식)
        - ordering: 정렬 필드

        Example: /api/production/work-orders/export/excel?fields=id,order_number,status
        """
        queryset = self.get_queryset()

        # 필터 적용
        queryset = self.filter_queryset(queryset)

        # 정렬 적용
        ordering = request.query_params.get('ordering')
        if ordering:
            queryset = queryset.order_by(ordering)

        # 필드 제한
        fields_param = request.query_params.get('fields')
        if fields_param:
            fields = fields_param.split(',')
        else:
            fields = self.get_export_fields()

        # 필드 이름 매핑
        field_names = self.get_export_field_names()

        # 엑셀 워크북 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.queryset.model._meta.verbose_name or 'Data'

        # 헤더 작성
        headers = [field_names.get(f, f) for f in fields]
        ws.append(headers)

        # 데이터 작성
        for item in queryset:
            serializer = self.get_serializer(item)
            row = [serializer.data.get(f, '') for f in fields]
            ws.append(row)

        # 응답 생성
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'{self.queryset.model._meta.model_name}_{timestamp}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # 엑셀 파일 저장
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        response.write(excel_buffer.getvalue())

        return response

    @action(detail=False, methods=['get'], url_path='export/json')
    def export_json(self, request):
        """
        JSON 형식으로 데이터 내보내기

        Query Parameters:
        - fields: 선택적 필드 목록 (쉼표로 구분)
        - filters: 필터 조건 (JSON 형식)
        - ordering: 정렬 필드

        Example: /api/production/work-orders/export/json?fields=id,order_number,status
        """
        queryset = self.get_queryset()

        # 필터 적용
        queryset = self.filter_queryset(queryset)

        # 정렬 적용
        ordering = request.query_params.get('ordering')
        if ordering:
            queryset = queryset.order_by(ordering)

        # 필드 제한
        fields_param = request.query_params.get('fields')
        serializer = self.get_serializer(queryset, many=True)

        data = serializer.data
        if fields_param:
            fields = fields_param.split(',')
            data = [{k: v for k, v in item.items() if k in fields} for item in data]

        return Response({
            'count': len(data),
            'results': data
        })

    @action(detail=False, methods=['post'], url_path='import/csv')
    def import_csv(self, request):
        """
        CSV 형식으로 데이터 가져오기

        Request Body (multipart/form-data):
        - file: CSV 파일

        Response:
        {
            "imported": int,
            "failed": int,
            "results": [...],
            "errors": [...]
        }
        """
        file = request.FILES.get('file')
        if not file:
            return Response(
                {'error': 'CSV 파일이 필요합니다.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # CSV 파일 읽기
        decoded_file = file.read().decode('utf-8')
        io_string = io.StringIO(decoded_file)
        reader = csv.DictReader(io_string)

        results = []
        errors = []
        imported_count = 0
        failed_count = 0

        for row in reader:
            serializer = self.get_serializer(data=row)
            if serializer.is_valid():
                try:
                    serializer.save()
                    results.append(serializer.data)
                    imported_count += 1
                except Exception as e:
                    errors.append({
                        'row': row,
                        'error': str(e)
                    })
                    failed_count += 1
            else:
                errors.append({
                    'row': row,
                    'error': serializer.errors
                })
                failed_count += 1

        return Response({
            'imported': imported_count,
            'failed': failed_count,
            'results': results,
            'errors': errors
        })

    @action(detail=False, methods=['post'], url_path='import/excel')
    def import_excel(self, request):
        """
        Excel 형식으로 데이터 가져오기

        Request Body (multipart/form-data):
        - file: Excel 파일 (.xlsx)

        Response:
        {
            "imported": int,
            "failed": int,
            "results": [...],
            "errors": [...]
        }
        """
        file = request.FILES.get('file')
        if not file:
            return Response(
                {'error': 'Excel 파일이 필요합니다.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active

            # 헤더 읽기
            headers = [cell.value for cell in ws[1]]

            # 데이터 읽기
            results = []
            errors = []
            imported_count = 0
            failed_count = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))
                serializer = self.get_serializer(data=row_data)
                if serializer.is_valid():
                    try:
                        serializer.save()
                        results.append(serializer.data)
                        imported_count += 1
                    except Exception as e:
                        errors.append({
                            'row': row_data,
                            'error': str(e)
                        })
                        failed_count += 1
                else:
                    errors.append({
                        'row': row_data,
                        'error': serializer.errors
                    })
                    failed_count += 1

            return Response({
                'imported': imported_count,
                'failed': failed_count,
                'results': results,
                'errors': errors
            })
        except Exception as e:
            return Response(
                {'error': f'Excel 파일 처리 실패: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['post'], url_path='import/json')
    def import_json(self, request):
        """
        JSON 형식으로 데이터 가져오기

        Request Body:
        {
            "items": [
                {field1: value1, field2: value2, ...},
                {field1: value1, field2: value2, ...},
                ...
            ]
        }

        Response:
        {
            "imported": int,
            "failed": int,
            "results": [...],
            "errors": [...]
        }
        """
        items = request.data.get('items', [])
        if not items:
            return Response(
                {'error': 'items 배열이 필요합니다.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        results = []
        errors = []
        imported_count = 0
        failed_count = 0

        for item in items:
            serializer = self.get_serializer(data=item)
            if serializer.is_valid():
                try:
                    serializer.save()
                    results.append(serializer.data)
                    imported_count += 1
                except Exception as e:
                    errors.append({
                        'item': item,
                        'error': str(e)
                    })
                    failed_count += 1
            else:
                errors.append({
                    'item': item,
                    'error': serializer.errors
                })
                failed_count += 1

        return Response({
            'imported': imported_count,
            'failed': failed_count,
            'results': results,
            'errors': errors
        })
